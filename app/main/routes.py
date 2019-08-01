import requests
import json
import webbrowser
import datetime
import pandas as pd
from urllib.parse import urlparse, parse_qs, urljoin
from flask import flash, render_template, redirect, url_for, jsonify
from app.main import bp
from app.main.forms import OAuthForm, TokenForm, RedirectForm, TestForm
from app.auth.forms import LoginForm
from flask import session
from flask import request
from app.utils import OAuthSignIn
from datetime import date, timedelta
from collections import OrderedDict
from flask_login import login_required
from openpyxl import load_workbook
from app import db
from app.models import Vat_return




def get_first_day(dt, d_years=0, d_months=0):
    # d_years, d_months are "deltas" to apply to dt
    y, m = dt.year + d_years, dt.month + d_months
    a, m = divmod(m-1, 12)
    return date(y+a, m+1, 1)

def get_last_day(dt):
    return get_first_day(dt, 0, 1) + timedelta(-1)

def get_first_tax_period():
    d = date.today()
    y = d.year
    return date(y, 4, 1)



@bp.route('/home')
@login_required
def home():

    org = session['org']
    title = 'Main Home'
    return render_template('main/home.html', title=title, org=org)



@bp.route('/authorize/<provider>')
@login_required
def oauth_authorize(provider):
    #if not current_user.is_anonymous():
    #    return redirect(url_for('home'))
    oauth = OAuthSignIn.get_provider(provider)
    return oauth.authorize()



@bp.route('/callback/<provider>')
@login_required
def oauth_callback(provider):
    #if not current_user.is_anonymous():
    #    return redirect(url_for('home'))
    oauth = OAuthSignIn.get_provider(provider)
    access_token = oauth.callback()
    
    return render_template('main/oauth_callback.html', title='OAuth-callback', access_token=access_token)



@bp.route('/refresh/<provider>')
@login_required
def oauth_refresh(provider):
    #if not current_user.is_anonymous():
    #    return redirect(url_for('home'))
    oauth = OAuthSignIn.get_provider(provider)
    access_token = oauth.refresh_token()
    
    return render_template('main/oauth_callback.html', title='OAuth-callback', access_token=access_token)



@bp.route('/retrieve/vat_obligations', methods=['GET','POST'])
@login_required
def vat_obligations():
    
    dt_from = request.form.get('datepicker1', get_first_tax_period().__format__('%Y-%m-%d'))
    dt_to = request.form.get('datepicker2', get_last_day(date.today()).__format__('%Y-%m-%d'))

    params = {
               'from': dt_from,
               'to': dt_to
             }

    
    vrn = session['vrn']  #'789904852'
    endpoint = '/organisations/vat/{}/obligations'.format(vrn)
    oauth = OAuthSignIn.get_provider('hmrc')
    response = oauth.get_data(endpoint, params)

    r = response.json()
    
    
    colnames = {'1':'start','2':'end','3':'due','4':'status','5':'received'}
    orderedcols = OrderedDict()

    for i in sorted(colnames):
        orderedcols[i] = colnames[i]


    if response.status_code != 200:
              
            flash('{}: - {}'.format(response.status_code, r['message']))
            
            return render_template('main/vat_obligations.html', title="vat obligations")


    d = r['obligations']
    

    return render_template('main/vat_obligations.html', title="vat obligations", my_data=d, colnames=orderedcols)




@bp.route('/retrieve/vat_liabilities', methods=['GET','POST'])
@login_required
def vat_liabilities():
    
    dt_from = request.form.get('datepicker1', get_first_tax_period().__format__('%Y-%m-%d'))
    dt_to = request.form.get('datepicker2', get_last_day(date.today()).__format__('%Y-%m-%d'))

    params = {
               'from': dt_from,
               'to': dt_to
             }

    
    vrn = session['vrn'] #'789904852'
    endpoint = '/organisations/vat/{}/liabilities'.format(vrn)
    oauth = OAuthSignIn.get_provider('hmrc')
    response = oauth.get_data(endpoint, params)

    r = response.json()

    colnames = {'1':'taxPeriod-from', '2':'taxPeriod-to','3':'type','4':'originalAmount','5':'oustandingAmount','6':'due'}
    orderedcols = OrderedDict()

    for i in sorted(colnames):
        orderedcols[i] = colnames[i]

    if response.status_code != 200:
              
            flash('{}: - {}'.format(response.status_code, r['message']))
            
            return render_template('main/vat_liabilities.html', title="vat liabilities")
    

    d = r['liabilities']
    
    
    return render_template('main/vat_liabilities.html', title="vat liabilities", my_data=d, colnames=orderedcols)





@bp.route('/retrieve/vat_payments', methods=['GET','POST'])
@login_required
def vat_payments():
    
    dt_from = request.form.get('datepicker1', get_first_tax_period().__format__('%Y-%m-%d'))
    dt_to = request.form.get('datepicker2', get_last_day(date.today()).__format__('%Y-%m-%d'))

    params = {
               'from': dt_from,
               'to': dt_to
             }

    
    vrn = session['vrn'] #'789904852'
    endpoint = '/organisations/vat/{}/payments'.format(vrn)
    oauth = OAuthSignIn.get_provider('hmrc')
    response = oauth.get_data(endpoint, params)

    r = response.json()
        
    
    colnames = {'1':'amount', '2':'received'}
    orderedcols = OrderedDict()

    for i in sorted(colnames):
        orderedcols[i] = colnames[i]
    
    if response.status_code != 200:
              
            flash('{}: - {}'.format(response.status_code, r['message']))
            
            return render_template('main/vat_payments.html', title="vat payments")
    

    d = r['payments']


    return render_template('main/vat_payments.html', title="vat payments", my_data=d, colnames=orderedcols)



@bp.route('/upload/vat_return/<string:period>', methods=['GET','POST'])
@login_required
def upload_return(period):
    
    rtn = { "periodKey": period }
    

    rows = { 1: {'id': 'periodKey', 'desc':'Box 0. Period Key'},
                 2: {'id': 'vatDueSales', 'desc':'Box 1. Vat Due on Sales'},
                 3: {'id': 'vatDueAcquisitions', 'desc':'Box 2. VAT Due on Acquisitions'},
                 4: {'id': 'totalVatDue', 'desc':'Box 3. Total VAT Due'},
                 5: {'id': 'vatReclaimedCurrPeriod', 'desc':'Box 4. VAT Reclaimed in Current Period'},
                 6: {'id': 'netVatDue', 'desc':'Box 5. Net VAT to be paid to HMRC or reclaimed'},
                 7: {'id': 'totalValueSalesExVAT', 'desc':'Box 6. Total Value of Sales (excl. VAT)'},
                 8: {'id': 'totalValuePurchasesExVAT', 'desc':'Box 7. Total value of Purchases (excl. VAT)'},
                 9: {'id': 'totalValueGoodsSuppliedExVAT', 'desc':'Box 8. Total Value of Goods Supplied (excl. VAT)'},
                 10:{'id': 'totalAcquisitionsExVAT', 'desc':'Box 9. Total Acquisitions (excl. VAT)'},
                 11:{'id':'finalised', 'desc':'Box 10. Finalised VAT Return'}
                 }

    orderedrows = OrderedDict()

    for i in sorted(rows):
        orderedrows[i] = rows[i]

    title = 'MTD File Upload'

    if request.method == 'POST':
        
        
        session['periodKey'] = period
        rtn = { "periodKey": period }


        # check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)

        f = request.files['file']

        wb = load_workbook(filename= f, read_only=True)
        ws = wb['Sheet1']

        # Read the cell values into a list of lists
        data_rows = []
        
        for row in ws['B7':'C16']:
            data_cols = []
            
            for cell in row:
                data_cols.append(cell.value)
                data_rows.append(data_cols)

       # Transform into dataframe
       
        df = pd.DataFrame(data_rows, columns=['key','value'])

       # Convert dataframe into dictionary
        ds = df.set_index('key')['value'].to_dict()
        

        rtn.update(ds)

        vr = Vat_return.query.filter_by(period_key=rtn['periodKey'], organisation_id = session['org']).first()
        if vr is None:
            vr = Vat_return(
                  organisation_id = session['org'],
                  period_key = rtn['periodKey'],
                  vat_due_sales = rtn['vatDueSales'],
                  vat_due_acquisitions = rtn['vatDueAcquisitions'],
                  total_vat_due = rtn['totalVatDue'],
                  vat_reclaimed_curr_period = rtn['vatReclaimedCurrPeriod'],
                  net_vat_due = rtn['netVatDue'],
                  total_value_sales_ex_vat = rtn['totalValueSalesExVAT'],
                  total_value_purchases_ex_vat = rtn['totalValuePurchasesExVAT'],
                  total_value_goods_supplied_ex_vat = rtn['totalValueGoodsSuppliedExVAT'],
                  total_acquisitions_ex_vat = rtn['totalAcquisitionsExVAT'],
                  finalised = True

                  )
        
            db.session.add(vr)
            db.session.commit()

            flash('VAT return for the selected period has been saved')


        else:
            vr.vat_due_sales = rtn['vatDueSales']
            vr.vat_due_acquisitions = rtn['vatDueAcquisitions']
            vr.total_vat_due = rtn['totalVatDue']
            vr.vat_reclaimed_curr_period = rtn['vatReclaimedCurrPeriod']
            vr.net_vat_due = rtn['netVatDue']
            vr.total_value_sales_ex_vat = rtn['totalValueSalesExVAT']
            vr.total_value_purchases_ex_vat = rtn['totalValuePurchasesExVAT']
            vr.total_value_goods_supplied_ex_vat = rtn['totalValueGoodsSuppliedExVAT']
            vr.total_acquisitions_ex_vat = rtn['totalAcquisitionsExVAT']
            
            db.session.commit()

            flash('The VAT return for the selected period has been updated')   

        return render_template('main/vat_upload.html', period=period, title=title, rtn = rtn, rows = orderedrows)
    
    
    return render_template('main/vat_upload.html', period=period, title=title, rtn = rtn, rows = orderedrows)




@bp.route('/submit/vat_return', methods=['GET','POST'])
@login_required
def submit_vat_return():

#   get periodKey for outstanding vat return.
    org = session['org'] 
    periodKey = session['periodKey']

    vrn = session['vrn'] #'789904852'
    endpoint = '/organisations/vat/{}/returns'.format(vrn)
    #url = urljoin('https://api.service.hmrc.gov.uk', endpoint)

    ''' 
    headers = {"Accept":"application/vnd.hmrc.1.0+json",
                "Content-Type": "application/json",
                "Authorization": 'Bearer %s' % session['access_token']
            }
    '''
    vat_return = Vat_return.query.filter_by(organisation_id = org).first()
    
    ''' check if vat return available for open period or if already submitted  '''
    
      
    if vat_return is not None:
        data = {'periodKey': vat_return.period_key,
                'vatDueSales': vat_return.vat_due_sales,
                'vatDueAcquisitions': vat_return.vat_due_acquisitions,
                'totalVatDue': vat_return.total_vat_due,
                'vatReclaimedCurrPeriod': vat_return.vat_reclaimed_curr_period,
                'netVatDue': vat_return.net_vat_due,
                'totalValueSalesExVAT': vat_return.total_value_sales_ex_vat,
                'totalValuePurchasesExVAT': vat_return.total_value_purchases_ex_vat,
                'totalValueGoodsSuppliedExVAT': vat_return.total_value_goods_supplied_ex_vat,
                'totalAcquisitionsExVAT': vat_return.total_acquisitions_ex_vat,
                'finalised': vat_return.finalised
                }

        
        
        oauth = OAuthSignIn.get_provider('hmrc')
        response = oauth.post_data(endpoint, data)
        '''
        response = requests.post(url, json=data, headers=headers)
        '''
        r=response.json()
        
        '''
        test = {'processingDate': '2019-07-25T17:01:16.837Z', 
                'paymentIndicator': 'DD', 
                'formBundleNumber': '603756869697', 
                'chargeRefNumber': 'U1IUpFzKlhpFrlVX'}
        '''

        if response.status_code != 200:
            err = r['errors']

            return render_template('main/vat_submission.html' )
        
        #vat_receipt = Vat_receipt(organisation_id = session['ord_id'], period_key) 
        msg = "HMRC has accepted your return"

    return render_template('main/vat_submission.html', response=r , msg=msg)
        
    
 


@bp.route('/view/vat_return/<string:period>', methods=['GET','POST'])
@login_required
def view_return(period):


#    
    periodKey = period
    vrn = session['vrn'] #'789904852'
    endpoint = '/organisations/vat/{}/returns/{}'.format(vrn, periodKey)
    oauth = OAuthSignIn.get_provider('hmrc')
    response = oauth.get_data(endpoint)

    r = response.json()
    
    if response.status_code != 200:
        err = r['errors']

        return render_template('main/view_return.html', errors=err )

    return render_template('main/view_return.html', response=r )







@bp.route('/test', methods=['GET','POST'])
def test():
    form = TestForm()
    new = 1
    url = 'http://docs.python.org/library/webbrowser.html'
    chrome_path = '"C:\Program Files (x86)\Google\Chrome\Application\Chrome.exe" %s'
    if form.validate_on_submit():
        webbrowser.get(chrome_path).open(url, new=new)

    return render_template('test.html', title='TEST', form=form)




'''
User ID
232362788506

Password
b8uyxmjfruaa

Corporation Tax UTR
1817122087

VAT Registration Number
205275977

Organisation Details
{"name":"Company KIVA0X","address":{"line1":"13 Mortimer Square","line2":"Warwick","postcode":"TS19 1PA"}}

'''